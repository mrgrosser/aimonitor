import csv
import hashlib
import hmac
import html
import io
import json
import os
import re
import sqlite3
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.governance import DB_PATH

MAX_IMPORT_BYTES = 12 * 1024 * 1024
ANONYMIZATION_KEY = (os.getenv("USAGE_ANONYMIZATION_KEY") or os.getenv("SESSION_SECRET") or "jo-ai-monitor-development-key").encode()


def init_usage_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with closing(sqlite3.connect(DB_PATH)) as db:
        db.execute("""CREATE TABLE IF NOT EXISTS usage_periods (
            period TEXT PRIMARY KEY, data_json TEXT NOT NULL, source_name TEXT NOT NULL,
            source_hash TEXT NOT NULL UNIQUE, imported_at TEXT NOT NULL, imported_by TEXT NOT NULL)""")
        db.commit()


def _number(value: Any, default: float = 0) -> float:
    if value is None or value == "": return default
    if isinstance(value, (int, float)): return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try: return float(cleaned)
    except ValueError: return default


def _integer(value: Any) -> int:
    return int(round(_number(value)))


def _currency_from_text(value: Any) -> float:
    match = re.search(r"\$\s*([0-9][0-9,]*(?:\.[0-9]+)?)", str(value or ""))
    return float(match.group(1).replace(",", "")) if match else 0


def _alias(value: Any) -> str:
    digest = hmac.new(ANONYMIZATION_KEY,str(value or "unknown").strip().lower().encode(),hashlib.sha256).hexdigest()[:8].upper()
    return f"User {digest}"


def _sheet_rows(workbook, name: str) -> list[list[Any]]:
    if name not in workbook.sheetnames: return []
    return [list(row) for row in workbook[name].iter_rows(values_only=True)]


def _find(rows: list[list[Any]], label: str, column: int = 0) -> list[Any] | None:
    target = label.strip().lower()
    return next((row for row in rows if len(row) > column and str(row[column] or "").strip().lower() == target), None)


def _table_after(rows: list[list[Any]], first_header: str, stop: str = "Total") -> list[list[Any]]:
    start = next((i for i, row in enumerate(rows) if row and str(row[0] or "").strip() == first_header), None)
    if start is None: return []
    result = []
    for row in rows[start + 1:]:
        if not row or row[0] is None: continue
        if str(row[0]).strip() == stop: break
        result.append(row)
    return result


def parse_xlsx(content: bytes, filename: str) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    summary_rows = _sheet_rows(workbook, "AI Usage Summary")
    product_rows = _sheet_rows(workbook, "Claude Product & Model")
    app_rows = _sheet_rows(workbook, "Copilot App Totals")
    copilot_user_rows = _sheet_rows(workbook, "Copilot User by App")
    claude_user_rows = _sheet_rows(workbook, "Claude User by Product")
    claude_detail_rows = _sheet_rows(workbook, "Claude Detail")
    if not summary_rows: raise ValueError("Missing required sheet: AI Usage Summary")

    heading = " ".join(str(x or "") for row in summary_rows[:4] for x in row)
    match = re.search(r"(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+20\d{2}", heading, re.I)
    period = match.group(0).title() if match else Path(filename).stem
    measure = {str(row[0]).strip():row for row in summary_rows if row and row[0]}
    users = measure.get("Users with recorded activity", [None, 0, 0])
    volume = measure.get("Total volume", [None, 0, 0])
    surfaces = measure.get("Distinct surfaces / products", [None, 0, 0])
    agents = measure.get("Agent-assisted volume", [None, 0, 0])
    spend = measure.get("July Claude usage spend", measure.get("Claude usage spend", [None, 0]))
    budget = measure.get("Claude monthly spend limit", [None, 0])
    purchased = measure.get("Claude Enterprise seats purchased", [None, 0])
    assigned = measure.get("Claude seats assigned", [None, 0])
    annual = _currency_from_text((measure.get("Claude Enterprise seats purchased") or [None, 0, ""])[2])

    products = []
    models = []
    section = None
    for row in product_rows:
        label = str(row[0] or "").strip()
        if label == "Product": section = "product"; continue
        if label == "Model": section = "model"; continue
        if not label or label == "By model": continue
        if label == "Total": continue
        if section == "product": products.append({"name":label,"requests":_integer(row[1]),"spend":round(_number(row[2]),2)})
        elif section == "model": models.append({"name":label,"requests":_integer(row[1]),"spend":round(_number(row[2]),2)})

    apps = [{"name":str(row[0]),"interactions":_integer(row[1]),"users":_integer(row[2])}
            for row in _table_after(app_rows,"App") if row[0] and str(row[0]) != "Total"]

    top_users = []
    copilot_header = _find(copilot_user_rows,"User") or []
    copilot_total_col = next((i for i,x in enumerate(copilot_header) if x == "Total"),17)
    for row in _table_after(copilot_user_rows,"User")[:10]:
        if len(row) > copilot_total_col:
            breakdown=[{"name":str(copilot_header[i]),"volume":_integer(row[i]),"spend":None,
                "spend_label":"Included in license fee","basis":"Provider-reported interactions"}
                for i in range(1,copilot_total_col) if i < len(row) and _integer(row[i])]
            top_users.append({"user":str(row[0]),"alias":_alias(row[0]),"provider":"Microsoft 365 Copilot",
                "volume":_integer(row[copilot_total_col]),"surfaces":len(breakdown),
                "active_days":_integer(row[copilot_total_col+2] if len(row)>copilot_total_col+2 else 0),
                "spend":None,"products":[x["name"] for x in breakdown],"breakdown":breakdown,
                "breakdown_note":"Exact July interactions by host app. Copilot has no incremental per-use charge in this source."})
    claude_matrix = {str(row[0]).strip().lower():row for row in _table_after(claude_user_rows,"User")}
    claude_header = _find(claude_user_rows,"User") or []
    claude_total_col = next((i for i,x in enumerate(claude_header) if x == "Total requests"),7)
    for row in claude_detail_rows:
        if row and row[0] == "by_user" and len(top_users) < 16:
            identity=str(row[1]); matrix=claude_matrix.get(identity.strip().lower(),[])
            matrix_total=_integer(matrix[claude_total_col]) if len(matrix)>claude_total_col else 0
            user_spend=round(_number(row[3]),2); breakdown=[]
            for i in range(1,min(claude_total_col,len(matrix))):
                requests=_integer(matrix[i])
                if requests:
                    breakdown.append({"name":str(claude_header[i]),"volume":requests,
                        "spend":round(user_spend*requests/matrix_total,2) if matrix_total else None,
                        "basis":"Estimated allocation using product request share"})
            top_users.append({"user":identity,"alias":_alias(identity),"provider":"Claude Enterprise",
                "volume":_integer(row[5]),"surfaces":len(breakdown),"active_days":0,"spend":user_spend,
                "products":[x["name"] for x in breakdown],"breakdown":breakdown,
                "breakdown_note":"User spend is July provider data. Product requests cover 26 May–24 Aug; product dollars are an estimated proportional allocation, not provider-reported spend."})

    claude_spend = round(_number(spend[1] if len(spend) > 1 else 0),2)
    annual_cost = annual or 4800.0
    monthly_seat = round(annual_cost / 12,2)
    assigned_count = _integer(assigned[1] if len(assigned) > 1 else 0)
    purchased_count = _integer(purchased[1] if len(purchased) > 1 else 0)
    data = {"period":period,"source":f"Imported from {filename}","summary":{
        "copilot_active_users":_integer(users[1]),"claude_active_users":_integer(users[2]),
        "copilot_interactions":_integer(volume[1]),"claude_requests":_integer(volume[2]),
        "claude_usage_spend":claude_spend,"claude_usage_budget":round(_number(budget[1]),2),
        "agent_interactions":_integer(agents[1]),"distinct_surfaces":_integer(surfaces[1])},
        "licensing":{"claude_seats_purchased":purchased_count,"claude_seats_assigned":assigned_count,
            "claude_seats_unassigned":max(0,purchased_count-assigned_count),"claude_annual_seat_cost":annual_cost,
            "monthly_seat_equivalent":monthly_seat,"estimated_monthly_run_rate":round(monthly_seat+claude_spend,2),
            "copilot_usage_cost":"Included in license fee"},"claude_products":products,"claude_models":models,
        "copilot_apps":apps,"top_users":top_users,"caveats":[
            "Copilot interactions and Claude API requests are different units and must not be totaled or compared as equivalent effort.",
            "Agentic products may issue multiple API requests for one user action.",
            "Named identities are retained for approved role-restricted reporting and replaced with stable aliases for other viewers.",
            "Claude per-product user spend is an estimated allocation because the provider export does not include a user-by-product spend cross-tab.",
            "Usage is not an employee performance measure."]}
    return validate_usage(data)


def parse_csv(content: bytes, filename: str) -> dict[str, Any]:
    text = content.decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    if not rows or not {"category","name"}.issubset(rows[0]):
        raise ValueError("CSV requires category and name columns; see the normalized import format in README.md")
    period = next((r.get("period") for r in rows if r.get("period")), None) or Path(filename).stem
    data = {"period":period,"source":f"Imported from {filename}","summary":{},"licensing":{},
        "claude_products":[],"claude_models":[],"copilot_apps":[],"top_users":[],"caveats":[]}
    for row in rows:
        category = (row.get("category") or "").strip().lower()
        name = (row.get("name") or "").strip()
        if category == "summary": data["summary"][name] = _number(row.get("value"))
        elif category == "licensing": data["licensing"][name] = _number(row.get("value")) if row.get("value") else row.get("text","")
        elif category == "claude_product": data["claude_products"].append({"name":name,"requests":_integer(row.get("requests")),"spend":round(_number(row.get("spend")),2)})
        elif category == "claude_model": data["claude_models"].append({"name":name,"requests":_integer(row.get("requests")),"spend":round(_number(row.get("spend")),2)})
        elif category == "copilot_app": data["copilot_apps"].append({"name":name,"interactions":_integer(row.get("interactions")),"users":_integer(row.get("users"))})
        elif category == "top_user": data["top_users"].append({"user":name,"alias":_alias(name),"provider":row.get("provider") or "Unknown","volume":_integer(row.get("volume")),"surfaces":_integer(row.get("surfaces")),"active_days":_integer(row.get("active_days")),"spend":round(_number(row.get("spend")),2),"breakdown":[]})
        elif category == "caveat" and name: data["caveats"].append(name)
    return validate_usage(data)


def validate_usage(data: dict[str, Any]) -> dict[str, Any]:
    errors, warnings = [], []
    if not str(data.get("period") or "").strip(): errors.append("Reporting period is required")
    summary = data.get("summary") or {}
    for field in ("copilot_interactions","claude_requests","claude_usage_spend"):
        if field not in summary: warnings.append(f"Summary field is missing: {field}")
    product_total = sum(_integer(x.get("requests")) for x in data.get("claude_products",[]))
    model_total = sum(_integer(x.get("requests")) for x in data.get("claude_models",[]))
    expected = _integer(summary.get("claude_requests"))
    if expected and product_total and product_total != expected: warnings.append(f"Claude product requests ({product_total:,}) do not reconcile to summary ({expected:,})")
    if expected and model_total and model_total != expected: warnings.append(f"Claude model requests ({model_total:,}) do not reconcile to summary ({expected:,})")
    app_total = sum(_integer(x.get("interactions")) for x in data.get("copilot_apps",[]))
    copilot = _integer(summary.get("copilot_interactions"))
    if copilot and app_total and app_total != copilot: warnings.append(f"Copilot app interactions ({app_total:,}) do not reconcile to summary ({copilot:,})")
    if errors: raise ValueError("; ".join(errors))
    data["validation"] = {"valid":True,"warnings":warnings,"totals":{"claude_products":product_total,
        "claude_models":model_total,"copilot_apps":app_total}}
    return data


def parse_usage_file(content: bytes, filename: str) -> tuple[dict[str, Any], str]:
    if not content: raise ValueError("The uploaded file is empty")
    if len(content) > MAX_IMPORT_BYTES: raise ValueError("Import exceeds the 12 MB size limit")
    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".xlsx": data = parse_xlsx(content, filename)
        elif suffix == ".csv": data = parse_csv(content, filename)
        else: raise ValueError("Only .xlsx and .csv usage files are supported")
    except ValueError: raise
    except Exception as exc: raise ValueError(f"Unable to parse {suffix or 'uploaded'} usage file") from exc
    return data, hashlib.sha256(content).hexdigest()


def save_usage_period(data: dict[str, Any], source_name: str, source_hash: str, actor: str, replace: bool = False) -> None:
    init_usage_db(); now = datetime.now(timezone.utc).isoformat(); period = str(data["period"])
    with closing(sqlite3.connect(DB_PATH)) as db:
        same = db.execute("SELECT period FROM usage_periods WHERE source_hash=?",(source_hash,)).fetchone()
        if same and same[0] != period: raise ValueError(f"This exact file was already imported as {same[0]}")
        exists = db.execute("SELECT 1 FROM usage_periods WHERE period=?",(period,)).fetchone()
        if exists and not replace: raise ValueError(f"{period} already exists; choose replace to overwrite it")
        db.execute("INSERT OR REPLACE INTO usage_periods(period,data_json,source_name,source_hash,imported_at,imported_by) VALUES(?,?,?,?,?,?)",
            (period,json.dumps(data,separators=(",",":")),source_name,source_hash,now,actor))
        db.commit()


def get_usage_period(period: str) -> dict[str, Any] | None:
    init_usage_db()
    with closing(sqlite3.connect(DB_PATH)) as db:
        row = db.execute("SELECT data_json,source_name,source_hash,imported_at,imported_by FROM usage_periods WHERE period=?",(period,)).fetchone()
    if not row: return None
    data=json.loads(row[0]); data["import"]={"source_name":row[1],"source_hash":row[2],"imported_at":row[3],"imported_by":row[4]}; return data


def list_usage_periods() -> list[dict[str, Any]]:
    init_usage_db()
    with closing(sqlite3.connect(DB_PATH)) as db:
        rows=db.execute("SELECT period,data_json,source_name,source_hash,imported_at,imported_by FROM usage_periods ORDER BY imported_at DESC").fetchall()
    result=[]
    for period,payload,name,digest,created,actor in rows:
        data=json.loads(payload); summary=data.get("summary",{})
        result.append({"period":period,"source_name":name,"source_hash":digest,"imported_at":created,"imported_by":actor,
            "copilot_interactions":summary.get("copilot_interactions",0),"claude_requests":summary.get("claude_requests",0),
            "claude_usage_spend":summary.get("claude_usage_spend",0)})
    return result


def usage_csv(data: dict[str, Any], actor: str = "", version: str = "") -> bytes:
    out=io.StringIO(); writer=csv.writer(out); writer.writerow(["JO AI Monitor Usage Report",data.get("period","")]); writer.writerow(["Generated at",datetime.now(timezone.utc).isoformat()]); writer.writerow(["Generated by",actor]); writer.writerow(["Version",version]); writer.writerow(["Source",data.get("source","")]); writer.writerow([])
    writer.writerow(["Summary metric","Value"])
    for key,value in (data.get("summary") or {}).items(): writer.writerow([key,value])
    writer.writerow([]); writer.writerow(["Claude product","Requests","Spend USD"])
    for row in data.get("claude_products",[]): writer.writerow([row.get("name"),row.get("requests"),row.get("spend")])
    writer.writerow([]); writer.writerow(["Claude model","Requests","Spend USD"])
    for row in data.get("claude_models",[]): writer.writerow([row.get("name"),row.get("requests"),row.get("spend")])
    writer.writerow([]); writer.writerow(["Copilot application","Interactions","Users"])
    for row in data.get("copilot_apps",[]): writer.writerow([row.get("name"),row.get("interactions"),row.get("users")])
    return out.getvalue().encode("utf-8-sig")


def usage_pdf(data: dict[str, Any], actor: str, version: str) -> bytes:
    stream=io.BytesIO(); styles=getSampleStyleSheet(); doc=SimpleDocTemplate(stream,pagesize=letter,rightMargin=.55*inch,leftMargin=.55*inch,topMargin=.55*inch,bottomMargin=.55*inch)
    story=[Paragraph("JO AI Monitor - Executive Usage & Spend Report",styles["Title"]),Paragraph(html.escape(str(data.get("period",""))),styles["Heading2"]),
        Paragraph(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} by {html.escape(actor)} - JO AI Monitor {html.escape(version)}",styles["BodyText"]),Spacer(1,12)]
    summary=data.get("summary",{}); licensing=data.get("licensing",{})
    metrics=[["Measure","Value"],["Copilot active users",summary.get("copilot_active_users",0)],["Copilot interactions",summary.get("copilot_interactions",0)],
        ["Claude active users",summary.get("claude_active_users",0)],["Claude API requests",summary.get("claude_requests",0)],
        ["Claude usage spend",f"${_number(summary.get('claude_usage_spend')):,.2f}"],["Estimated monthly run rate",f"${_number(licensing.get('estimated_monthly_run_rate')):,.2f}"],
        ["Unassigned Claude seats",licensing.get("claude_seats_unassigned",0)]]
    table=Table(metrics,colWidths=[3.6*inch,2.6*inch]); table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#172231")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#D8DEE6")),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F6F8FA")]),("PADDING",(0,0),(-1,-1),7)])); story += [table,Spacer(1,14)]
    for title,rows,columns in [("Claude products",data.get("claude_products",[]),[("Product","name"),("Requests","requests"),("Spend","spend")]),
            ("Claude models",data.get("claude_models",[]),[("Model","name"),("Requests","requests"),("Spend","spend")]),
            ("Copilot applications",data.get("copilot_apps",[]),[("Application","name"),("Interactions","interactions"),("Users","users")])]:
        if title == "Copilot applications": story.append(PageBreak())
        story.append(Paragraph(title,styles["Heading2"])); values=[[x[0] for x in columns]]+[[f"${_number(row.get(key)):,.2f}" if key == "spend" else row.get(key,"") for _,key in columns] for row in rows]
        t=Table(values,repeatRows=1,colWidths=[3.5*inch,1.35*inch,1.35*inch]); t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#315FDC")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#D8DEE6")),("PADDING",(0,0),(-1,-1),6)])); story += [t,Spacer(1,12)]
    story.append(Paragraph("Governance notes",styles["Heading2"]));
    for note in data.get("caveats",[]): story.append(Paragraph(f"- {html.escape(str(note))}",styles["BodyText"]))
    def footer(canvas, document):
        canvas.saveState(); canvas.setFont("Helvetica",8); canvas.setFillColor(colors.HexColor("#697386"));
        canvas.drawString(.55*inch,.32*inch,f"JO AI Monitor {version}"); canvas.drawRightString(7.95*inch,.32*inch,f"Page {document.page}"); canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer); return stream.getvalue()


def usage_html(data: dict[str, Any], actor: str, version: str) -> str:
    s=data.get("summary",{}); l=data.get("licensing",{})
    def rows(items,cols): return "".join("<tr>"+"".join(f"<td>{html.escape(str(x.get(c,'')))}</td>" for c in cols)+"</tr>" for x in items)
    return f"""<!doctype html><html><head><meta charset='utf-8'><title>JO AI Monitor Report</title><style>body{{font:14px Arial;color:#172231;max-width:1000px;margin:36px auto;padding:0 20px}}h1{{margin-bottom:2px}}.meta{{color:#697386}}.cards{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:24px 0}}.card{{border:1px solid #dce1e6;padding:14px;border-radius:8px}}.card b{{display:block;font-size:24px;margin-top:6px}}table{{width:100%;border-collapse:collapse;margin:8px 0 24px}}th,td{{border-bottom:1px solid #e4e7eb;padding:8px;text-align:left}}th{{background:#172231;color:#fff}}@media print{{button{{display:none}}body{{margin:0}}}}</style></head><body><button onclick='print()'>Print / Save PDF</button><h1>JO AI Monitor</h1><h2>Executive Usage &amp; Spend Report</h2><p class='meta'>{html.escape(str(data.get('period','')))} · Generated by {html.escape(actor)} · {html.escape(version)}</p><div class='cards'><div class='card'>Copilot interactions<b>{_integer(s.get('copilot_interactions')):,}</b></div><div class='card'>Claude requests<b>{_integer(s.get('claude_requests')):,}</b></div><div class='card'>Claude spend<b>${_number(s.get('claude_usage_spend')):,.2f}</b></div><div class='card'>Monthly run rate<b>${_number(l.get('estimated_monthly_run_rate')):,.2f}</b></div></div><h2>Claude products</h2><table><tr><th>Product</th><th>Requests</th><th>Spend</th></tr>{rows(data.get('claude_products',[]),['name','requests','spend'])}</table><h2>Claude models</h2><table><tr><th>Model</th><th>Requests</th><th>Spend</th></tr>{rows(data.get('claude_models',[]),['name','requests','spend'])}</table><h2>Copilot applications</h2><table><tr><th>Application</th><th>Interactions</th><th>Users</th></tr>{rows(data.get('copilot_apps',[]),['name','interactions','users'])}</table></body></html>"""
